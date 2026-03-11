"""
Excel Writer Tool - Write user stories to Excel backlog

This tool appends new user stories to the Excel backlog file.
"""

import os
from typing import List
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from graph.state import GraphState, UserStory
import config


def init_excel_file(file_path: str, sheet_name: str = "UserStories"):
    """
    Initialize a new Excel file with headers
    
    Args:
        file_path: Path to create the Excel file
        sheet_name: Name of the sheet
    """
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Define headers
    headers = [
        "Story ID", "Title", "Description", "Acceptance Criteria",
        "Priority", "Status", "Related Requirement", "Related CR"
    ]
    
    # Write headers with styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Adjust column widths
    column_widths = [15, 40, 60, 80, 12, 25, 12, 25]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Save the file
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
    print(f"Initialized Excel file: {file_path}")


def append_stories_to_excel(
    stories: List[UserStory],
    file_path: str,
    sheet_name: str = "UserStories",
    related_requirement: str = None,
    related_cr: str = None
) -> bool:
    """
    Append user stories to the Excel backlog
    
    Args:
        stories: List of user stories to add
        file_path: Path to the Excel file
        sheet_name: Name of the sheet
        related_requirement: Related requirement ID
        related_cr: Related change request ID
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Check if file exists, if not create it
        if not os.path.exists(file_path):
            init_excel_file(file_path, sheet_name)
        
        # Load existing workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        
        # Find the last row with data
        last_row = ws.max_row + 1
        
        # Write each story
        for story in stories:
            ws.cell(row=last_row, column=1, value=story.get("story_id", ""))
            ws.cell(row=last_row, column=2, value=story.get("title", ""))
            ws.cell(row=last_row, column=3, value=story.get("description", ""))
            
            # Format acceptance criteria as pipe-separated string
            ac = story.get("acceptance_criteria", [])
            ac_text = " | ".join(ac) if isinstance(ac, list) else str(ac)
            ws.cell(row=last_row, column=4, value=ac_text)
            
            ws.cell(row=last_row, column=5, value=story.get("priority", "Medium"))
            ws.cell(row=last_row, column=6, value=story.get("status", "To Do"))
            ws.cell(row=last_row, column=7, value=related_requirement or story.get("related_requirement", ""))
            ws.cell(row=last_row, column=8, value=related_cr or story.get("related_cr", ""))
            
            last_row += 1
        
        # Save the workbook
        wb.save(file_path)
        print(f"Successfully added {len(stories)} stories to {file_path}")
        return True
        
    except Exception as e:
        print(f"Error writing to Excel: {e}")
        return False


def excel_writer_node(state: GraphState) -> GraphState:
    """
    Excel Writer Node
    
    Writes generated user stories to the Excel backlog
    
    Args:
        state: Current graph state with generated_stories
        
    Returns:
        Updated state with final status
    """
    generated_stories = state.get("generated_stories", [])
    email_type = state.get("email_type")
    requirement_doc = state.get("requirement_doc")
    change_request_doc = state.get("change_request_doc")
    
    if not generated_stories:
        return {
            **state,
            "status": "error",
            "message": "No stories to write to Excel"
        }
    
    try:
        # Get related IDs
        related_requirement = None
        related_cr = None
        
        if requirement_doc:
            related_requirement = requirement_doc.get("doc_id")
        if change_request_doc:
            related_cr = change_request_doc.get("cr_id")
        
        # Write to Excel
        success = append_stories_to_excel(
            stories=generated_stories,
            file_path=config.BACKLOG_FILE_PATH,
            sheet_name=config.BACKLOG_SHEET_NAME,
            related_requirement=related_requirement,
            related_cr=related_cr
        )
        
        if success:
            return {
                **state,
                "status": "success",
                "message": f"Successfully wrote {len(generated_stories)} stories to backlog"
            }
        else:
            return {
                **state,
                "status": "error",
                "message": "Failed to write stories to Excel"
            }
            
    except Exception as e:
        return {
            **state,
            "status": "error",
            "message": f"Excel writer error: {str(e)}"
        }


# For testing purposes
if __name__ == "__main__":
    # Test the Excel writer
    test_stories = [
        {
            "story_id": "STORY-001",
            "title": "Add Supplier",
            "description": "As a user, I want to add new suppliers",
            "acceptance_criteria": ["Form with validation", "Save to database"],
            "priority": "High",
            "status": "To Do"
        },
        {
            "story_id": "STORY-002",
            "title": "Edit Supplier",
            "description": "As a user, I want to edit supplier information",
            "acceptance_criteria": ["Pre-filled form", "Update database"],
            "priority": "High",
            "status": "To Do"
        }
    ]
    
    # Test append
    success = append_stories_to_excel(
        test_stories,
        "data/test_backlog.xlsx",
        "UserStories"
    )
    print("Write successful:", success)

